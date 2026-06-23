"""Excel-backed user authentication.

Users are maintained manually in an .xlsx file with columns:
    id, password, name, role        (role = 'admin' or 'user')

The app only ever READS this file (never writes), so adding/removing users and
setting who is an admin is done by editing the spreadsheet. It's re-read on every
login attempt, so manual edits take effect immediately without a restart.

Exam assignments for non-admin users are managed separately, in-app — see access.py.
"""
from __future__ import annotations

import openpyxl

from .config import USERS_XLSX

_HEADERS = ("id", "password", "name", "role")


def ensure_users_file() -> None:
    """Create users.xlsx (with a default admin) if missing; add `role` if absent."""
    if not USERS_XLSX.exists():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "users"
        ws.append(list(_HEADERS))
        ws.append(["admin", "admin123", "Administrator", "admin"])  # default seed
        wb.save(USERS_XLSX)
        return
    _migrate_add_role_column()


def _migrate_add_role_column() -> None:
    """If an existing file predates the `role` column, append it (admin id -> admin).

    Best-effort: if the file is open/locked (PermissionError), skip — load_users()
    falls back to treating the `admin` id as an admin when no role column exists.
    """
    try:
        wb = openpyxl.load_workbook(USERS_XLSX)
    except Exception:
        return
    try:
        ws = wb.active
        headers = [str(c.value).strip().lower() if c.value is not None else "" for c in ws[1]]
        if "role" in headers:
            return
        role_col = len(headers) + 1
        ws.cell(row=1, column=role_col, value="role")
        id_col = headers.index("id") + 1 if "id" in headers else 1
        for row in range(2, ws.max_row + 1):
            uid = ws.cell(row=row, column=id_col).value
            if uid is None:
                continue
            ws.cell(row=row, column=role_col,
                    value="admin" if str(uid).strip().lower() == "admin" else "user")
        wb.save(USERS_XLSX)
    except PermissionError:
        pass  # file open in Excel — fall back to id-based admin detection
    finally:
        wb.close()


def load_users() -> dict[str, dict]:
    """Return {id: {id, password, name, role, is_admin}} read from the Excel file."""
    users: dict[str, dict] = {}
    if not USERS_XLSX.exists():
        return users
    wb = openpyxl.load_workbook(USERS_XLSX, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            return users
        idx = {str(h).strip().lower(): i for i, h in enumerate(headers) if h is not None}
        for row in rows:
            if not row:
                continue

            def cell(col: str) -> str:
                i = idx.get(col)
                if i is None or i >= len(row) or row[i] is None:
                    return ""
                return str(row[i]).strip()

            uid = cell("id")
            if not uid:
                continue
            # If the file has no `role` column yet, bootstrap the `admin` id as admin.
            if "role" in idx:
                role = (cell("role") or "user").lower()
            else:
                role = "admin" if uid.lower() == "admin" else "user"
            users[uid] = {
                "id": uid,
                "password": cell("password"),
                "name": cell("name") or uid,
                "role": role,
                "is_admin": role == "admin",
            }
    finally:
        wb.close()
    return users


def authenticate(user_id: str, password: str) -> dict | None:
    """Return the session user dict on a valid id+password match, else None."""
    user = load_users().get(str(user_id).strip())
    if user and password and user["password"] == password:
        return {"id": user["id"], "name": user["name"], "is_admin": user["is_admin"]}
    return None


# --- in-app user management (writes to the Excel file) ---------------------

class UsersFileLocked(RuntimeError):
    """Raised when users.xlsx can't be written (open in Excel)."""


def _open_for_write():
    ensure_users_file()
    try:
        return openpyxl.load_workbook(USERS_XLSX)
    except PermissionError:
        raise UsersFileLocked("users.xlsx is open in Excel — close it and try again.")


def _header_map(ws) -> list[str]:
    headers = [str(c.value).strip().lower() if c.value is not None else "" for c in ws[1]]
    for h in ("id", "password", "name", "role"):
        if h not in headers:
            headers.append(h)
            ws.cell(row=1, column=len(headers), value=h)
    return headers


def _save(wb) -> None:
    try:
        wb.save(USERS_XLSX)
    except PermissionError:
        raise UsersFileLocked("users.xlsx is open in Excel — close it and try again.")
    finally:
        wb.close()


def create_user(user_id: str, password: str, name: str, role: str = "user") -> dict:
    user_id = str(user_id).strip()
    if not user_id:
        raise ValueError("User ID is required.")
    if not password:
        raise ValueError("Password is required.")
    role = "admin" if str(role).strip().lower() == "admin" else "user"
    if user_id in load_users():
        raise ValueError(f"User '{user_id}' already exists.")
    wb = _open_for_write()
    ws = wb.active
    headers = _header_map(ws)
    vals = {"id": user_id, "password": str(password), "name": (name or user_id).strip() or user_id, "role": role}
    ws.append([vals.get(h, "") for h in headers])
    _save(wb)
    return {"id": user_id, "name": vals["name"], "role": role, "is_admin": role == "admin"}


def delete_user(user_id: str) -> None:
    user_id = str(user_id).strip()
    if user_id not in load_users():
        raise ValueError("User not found.")
    wb = _open_for_write()
    ws = wb.active
    headers = [str(c.value).strip().lower() if c.value is not None else "" for c in ws[1]]
    id_col = (headers.index("id") + 1) if "id" in headers else 1
    target = None
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=id_col).value
        if v is not None and str(v).strip() == user_id:
            target = r
            break
    if target is not None:
        ws.delete_rows(target, 1)
    _save(wb)
